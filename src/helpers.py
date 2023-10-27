import os
import glob
import yfinance as yf
import xlwings as xw
from openpyxl import load_workbook
from datetime import date
from src.ticker_symbols import ticker_symbols as ts, tickers as t


def get_today(*args):
    today_file = date.today().strftime(f'%d{args[0]}%m{args[0]}%Y')
    today_cell = date.today().strftime(f'%d{args[1]}%m{args[1]}%Y')
    return today_file, today_cell


def change_file_name(new_file_name):
    current_file = glob.glob('*.xlsx')[0]
    os.rename(current_file, new_file_name)


def get_cell(file_name):
    app = xw.App(visible=False)
    xlsx_book = app.books.open(file_name)
    sheet = xlsx_book.sheets['Portfolio']
    file_date = sheet['A1'].value
    total_before = round(sheet['C440'].value, 2)
    total_before_cash = round(sheet['C474'].value, 2)
    xlsx_book.close()
    return file_date, total_before, total_before_cash


def upload_data():
    tickers = yf.Tickers(t)
    data = tickers.tickers.items()
    return data


def get_values(data, ws):
    counter = 0
    report = []
    for key, value in data:
        stock_info = value.info
        symbol = stock_info['symbol']
        current_price = stock_info['currentPrice']
        currency = stock_info['currency']
        previous_close = stock_info['previousClose']
        counter += 1
        cell = ts.get(symbol)
        ws[cell] = current_price
        symbols = {'symbol': symbol, 'current_price': current_price,
                   'previous_close': previous_close, 'currency': currency}
        report.append(symbols)
    return report, counter


def get_values2(data, ws):
    counter = 0
    report = []
    pay_attention = []
    for elem in data:
        stock_info = elem[1].fast_info
        symbol = elem[0]
        current_price = stock_info['last_price']
        currency = stock_info['currency']
        prev_close = stock_info['regularMarketPreviousClose']
        pay_attention.append(symbol) if str(prev_close) == 'nan' else None
        previous_close = prev_close if str(prev_close) != 'nan' else stock_info['previousClose']
        counter += 1
        cell = ts.get(symbol)
        ws[cell] = current_price
        symbols = {'symbol': symbol, 'current_price': current_price,
                   'previous_close': previous_close, 'currency': currency}
        report.append(symbols)
    return report, counter, pay_attention


def write_data(new_file_name, today_cell, data):
    wb = load_workbook(new_file_name)
    ws = wb['Portfolio']
    ws['A1'] = today_cell
    report, counter = get_values(data, ws)
    # if the get_values(data, ws) function will not work, use get_values2(data, ws) function below
    # report, counter, pay_attention = get_values2(data, ws)
    wb.save(new_file_name)
    wb.close()
    return report, counter
    # return report, counter, pay_attention


def get_margin(cur_close, pre_close):
    if cur_close > pre_close:
        margin = ((cur_close - pre_close) / pre_close) * 100
        return round(margin, 2)
    elif cur_close < pre_close:
        margin = ((cur_close - pre_close) / pre_close) * 100
        return round(margin, 2)
    else:
        return '0'


def get_analytics(report_list, txt_edit, END):
    for elem in report_list:
        elem['margin'] = get_margin(elem['current_price'], elem['previous_close'])
        if int(elem['margin']) >= 5:
            text = f"{elem['symbol']}: {round(elem['current_price'], 2)}{elem['currency']} " \
                   f"({round(elem['previous_close'], 2)}{elem['currency']}) ↑ рост +{elem['margin']}%\n"
            txt_edit.insert(END, text)
        elif int(elem['margin']) <= -5:
            text = f"{elem['symbol']}: {round(elem['current_price'], 2)}{elem['currency']} " \
                   f"({round(elem['previous_close'], 2)}{elem['currency']}) ↓ падение {elem['margin']}%\n"
            txt_edit.insert(END, text)


def get_total(file_name, cell):
    app = xw.App(visible=False)
    xlsx_book = app.books.open(file_name)
    sheet = xlsx_book.sheets['Portfolio']
    total = sheet[cell].value
    xlsx_book.close()
    return round(total, 2)


def data_print(*args):
    args[6].insert(args[7], f'Total ({args[0]}) —> {args[1]}\n')
    args[6].insert(args[7], f'Total ({args[2]}) —> {args[3]}\n')
    args[6].insert(args[7], f'Разница —> {round(args[4], 2)}\n')
    args[6].insert(args[7], f'Total + Cash ({args[0]}) —> {round(args[5], 2)}\n')


def pay_attention_print(pay_attention, txt_edit, END):
    pay_attention_text = ', '.join(pay_attention)
    txt_edit.insert(END, f'ВНИМАНИЕ!: цена последнего закрытия {pay_attention_text} может быть не корректной!\n')
